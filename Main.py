import PautaParse as pp
from openpyxl import Workbook
from openpyxl import load_workbook
import os


current = os.getcwd()
path = "/pautaFinal/"

listaPautas =  os.listdir(current+path)

for pauta in listaPautas:
    nomeDisciplina = pauta.split(".")[0]

    complete = current + path + pauta

    turma = pp.open(complete)
    total =  turma["size"]

    wb = load_workbook("MODELO.xlsx")
    ws = wb.active

    ws["D6"] = turma["disciplina"]

    ws["D8"] = turma["prof"]
    for i in range(0, total):
         ws["E" + str(12 + i)] = turma["nome"][i]
         ws["G" + str(12 + i)] = turma["dre"][i]
    ws.print_area = 'B2:V48'
    wb.save(current + path + nomeDisciplina + ".xlsx")

__name__ = "main"